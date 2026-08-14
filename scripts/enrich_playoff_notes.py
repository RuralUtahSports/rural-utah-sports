import io
import json
import os
import re
import urllib.request
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

SHEET_ID = os.environ["SHEET_ID"]
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
ROOT = Path(__file__).resolve().parents[1]
PAGE_DATA_DIR = ROOT / "team-page-data"

ALIASES = {
    "GUNNISON": "GUNNISON VALLEY",
    "MAPLE MTN": "MAPLE MOUNTAIN",
    "MONUMENT VAL": "MONUMENT VALLEY",
    "CEDAR": "CEDAR CITY",
    "SUMMIT": "SUMMIT ACADEMY",
    "WASATCH ACAD": "WASATCH ACADEMY",
    "WASATCH ACAD.": "WASATCH ACADEMY",
    "HINKLEY": "HINCKLEY",
    "BY HIGH": "BYH",
    "BRIGHAM YOUNG": "BYH",
}


def clean(value):
    return str(value or "").strip()


def canonical_team(value):
    key = " ".join(clean(value).upper().split()).rstrip(".").strip()
    if key.startswith("WASATCH ACAD"):
        return "WASATCH ACADEMY"
    return ALIASES.get(key, key)


def slug(value):
    return re.sub(r"^-+|-+$", "", re.sub(r"[^a-z0-9]+", "-", canonical_team(value).lower()))


def format_date(value):
    if isinstance(value, (datetime, date)):
        return f"{value.month}/{value.day}/{value.year}"
    text = clean(value)
    for fmt in ("%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            return f"{parsed.month}/{parsed.day}/{parsed.year}"
        except ValueError:
            pass
    return text


def parse_date(value):
    try:
        return datetime.strptime(clean(value), "%m/%d/%Y")
    except ValueError:
        return None


def number(value):
    if value is None or value == "":
        return None
    try:
        n = float(value)
        return int(n) if n.is_integer() else n
    except (TypeError, ValueError):
        return None


def is_playoff_note(note):
    text = clean(note).upper()
    if not text:
        return False
    return bool(re.search(
        r"\b(?:PO|PLAYOFF|QF|QUARTERFINAL|QUARTERFINALS|SF|SEMIFINAL|SEMIFINALS|FINAL|FINALS|CHAMPIONSHIP|CHAMPIONSHIPS|RND|ROUND|RD|1ST|2ND|3RD)\b",
        text,
    ))


def result_from_row(value, team_score, opponent_score):
    result = clean(value).upper()
    if result in {"W", "L", "T"}:
        return result
    if team_score is None or opponent_score is None:
        return ""
    if team_score > opponent_score:
        return "W"
    if team_score < opponent_score:
        return "L"
    return "T"


def merge_game(existing, incoming):
    old_notes = clean(existing.get("notes"))
    new_notes = clean(incoming.get("notes"))
    if len(new_notes) > len(old_notes):
        existing["notes"] = new_notes
    existing["playoff"] = bool(existing.get("playoff") or incoming.get("playoff") or is_playoff_note(existing.get("notes")))
    return existing


def dedupe_schedules(schedules):
    exact_removed = 0
    near_removed = 0
    out = {}
    for year, games in (schedules or {}).items():
        exact = {}
        first_pass = []
        for game in games:
            key = (clean(game.get("date")), canonical_team(game.get("opponent")))
            game["opponent"] = key[1]
            if key in exact:
                exact_removed += 1
                merge_game(first_pass[exact[key]], game)
                continue
            exact[key] = len(first_pass)
            first_pass.append(game)

        ordered = sorted(enumerate(first_pass), key=lambda item: (parse_date(item[1].get("date")) or datetime.max, item[0]))
        drop = set()
        last_by_signature = {}
        for idx, game in ordered:
            game_date = parse_date(game.get("date"))
            if not game_date:
                continue
            signature = (
                canonical_team(game.get("opponent")),
                game.get("teamScore"),
                game.get("opponentScore"),
                game.get("result"),
            )
            prior = last_by_signature.get(signature)
            if prior:
                days = (game_date - prior[0]).days
                if 0 < days <= 3:
                    near_removed += 1
                    merge_game(first_pass[prior[1]], game)
                    drop.add(idx)
                    continue
            last_by_signature[signature] = (game_date, idx)
        out[str(year)] = [game for i, game in enumerate(first_pass) if i not in drop]
    return out, exact_removed, near_removed


def main():
    req = urllib.request.Request(EXPORT_URL, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=120) as response:
        workbook_bytes = response.read()

    wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet_lookup = {}
    for name in wb.sheetnames:
        sheet_lookup.setdefault(canonical_team(name), []).append(name)

    with open(ROOT / "teams-data.json", encoding="utf-8") as f:
        teams = json.load(f)

    enriched_teams = 0
    enriched_games = 0
    playoff_games = 0
    exact_removed = 0
    near_removed = 0

    for team_obj in teams:
        team = canonical_team(team_obj.get("team"))
        sheet_names = sheet_lookup.get(team, [])
        page_path = PAGE_DATA_DIR / f"{slug(team)}.json"
        if not sheet_names or not page_path.exists():
            continue

        schedules = {}
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, min_col=2, max_col=8, values_only=True):
                raw_date, raw_team_score, raw_opponent, raw_opponent_score, raw_notes, raw_result, raw_year = row
                game_date = format_date(raw_date)
                opponent = canonical_team(raw_opponent)
                team_score = number(raw_team_score)
                opponent_score = number(raw_opponent_score)
                notes = clean(raw_notes)

                if not game_date or not opponent or opponent in {"OPPONENT", "NONE"}:
                    continue
                if team_score is None or opponent_score is None:
                    continue

                year = number(raw_year)
                if year is None:
                    match = re.search(r"(\d{4})$", game_date)
                    year = int(match.group(1)) if match else None
                if year is None:
                    continue
                year = int(year)

                result = result_from_row(raw_result, team_score, opponent_score)
                playoff = is_playoff_note(notes)
                schedules.setdefault(str(year), []).append({
                    "date": game_date,
                    "opponent": opponent,
                    "teamScore": team_score,
                    "opponentScore": opponent_score,
                    "result": result,
                    "playoff": playoff,
                    "notes": notes,
                })

        schedules, removed_exact, removed_near = dedupe_schedules(schedules)
        exact_removed += removed_exact
        near_removed += removed_near
        team_games = sum(len(games) for games in schedules.values())
        team_playoffs = sum(1 for games in schedules.values() for game in games if game.get("playoff"))
        enriched_games += team_games
        playoff_games += team_playoffs

        if not schedules:
            continue

        with open(page_path, encoding="utf-8") as f:
            payload = json.load(f)
        payload["schedules"] = schedules
        with open(page_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, separators=(",", ":"))
        enriched_teams += 1

    print(f"Team schedule files enriched: {enriched_teams}")
    print(f"Schedule games enriched after dedupe: {enriched_games}")
    print(f"Playoff games identified: {playoff_games}")
    print(f"Schedule duplicates removed during enrichment: {exact_removed + near_removed} ({exact_removed} exact, {near_removed} near-date)")

    if enriched_teams == 0 or enriched_games == 0 or playoff_games == 0:
        raise RuntimeError("Playoff-note enrichment produced empty data; refusing to commit.")


if __name__ == "__main__":
    main()
